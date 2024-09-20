import os
from copy import copy

import pyodbc
import pandas as pd
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill, Font

from export import exportDataFrameEncaissement, exportDataFrameSimple, merge_excel_files
from utils import write_log, extract_values_in_json, get_today


def connexionSQlServer(server, base):
    conn = None
    value = f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={server};DATABASE={base};UID=reader;PWD=m1234"
    print(value)
    try:
        conn = pyodbc.connect(value)
    except pyodbc.Error as e:
        # Si une erreur se produit lors de la connexion, ajoutez le serveur à la liste des bases sans accès
        print("===============================")
        print(f"Erreur de Connexion pour {server} à la base {base} ")
        print("===============================")
        write_log(f"Erreur de connexion : {str(e)}")
    except Exception as e:
        write_log(f"Erreur Exception : {str(e)}")
    return conn


def get_column_index_by_title(sheet, title):
    for col in sheet.iter_cols(1, sheet.max_column):
        if col[0].value == title:
            return col[0].col_idx
    return 3


def getDataLink(connexion, dates=None):
    if connexion is not None:
        try:
            finds = extract_values_in_json('SQL')
            if dates is None:
                dates = get_today()
            resumer = {}
            details = None
            pal_1 = None
            pal_2 = None
            pal_3 = None
            for key, value in finds.items():
                if key not in ['details', 'palm_today', 'palm_month', 'palm_all']:
                    with connexion.cursor() as cursor:
                        query = str(value).replace(
                            "{today}", dates.strftime('%Y-%m-%d')
                        ).replace("{year}", dates.strftime('%Y')
                                  ).replace('{month}', dates.strftime('%m'))
                        cursor.execute(query)
                        rows = cursor.fetchall()
                        tuples = []
                        if rows:
                            tuples = [tuple(row) for row in rows]
                        resumer[key] = tuples

                elif key in ['details', ]:

                    with connexion.cursor() as cursor:
                        query = str(value).replace("{today}", dates.strftime('%Y-%m-%d'))
                        cursor.execute(query)
                        rows = cursor.fetchall()
                        rows = [tuple(row) for row in rows]
                        if all(isinstance(row, tuple) for row in rows):
                            details = pd.DataFrame(rows,
                                                   columns=["N° PIECE", "REF", "DESIGNATION", "QTE", "PRIX UNITAIRE",
                                                            "MONTANT TTC",
                                                            "MODE"])
                        else:
                            print("Pas un tuple")
                else:
                    with connexion.cursor() as cursor:
                        try:
                            query = str(value).replace("{today}", dates.strftime('%Y-%m-%d'))
                            cursor.execute(query)
                            rows = cursor.fetchall()
                            rows = [tuple(row) for row in rows]
                            if all(isinstance(row, tuple) for row in rows):
                                if str(key).replace("palm_", '') == 'today':
                                    pal_1 = pd.DataFrame(rows,
                                                         columns=["REF", "DESIGNATION", "QUANTITE", "MONTANT TTC"])

                                elif str(key).replace("palm_", '') == 'month':
                                    pal_2 = pd.DataFrame(rows,
                                                         columns=["REF", "DESIGNATION", "QUANTITE", "MONTANT TTC"])
                                elif str(key).replace("palm_", '') == 'all':
                                    pal_3 = pd.DataFrame(rows,
                                                         columns=["ARTICLE", "DESIGNATION", "QUANTITE", "MONTANT TTC"])

                        except Exception as e:
                            print(f"Erreur : {key}")
                            write_log(str(e))
                            pass

            # header_row = ['INTITULE', f'JOUR {dates.strftime('%d-%m-%Y')}', f"MOIS ENCOURS {dates.strftime('%B')}",
                        #   f"ANNEE ENCOURS {dates.strftime('%Y')}"]

            # excel_1 = exportDataFrameEncaissement(resumer, dates, header_row)
            excel_2 = exportDataFrameSimple(details, f"VENTE DU JOUR DETAIL {dates.strftime('%d-%m-%Y')}",
                                            [4, 5, 6])
     

            storage_dir = f'storage/{dates.strftime("%d-%m-%Y")}'
            if not os.path.exists(storage_dir):
                os.makedirs(storage_dir)

            # wb1 = load_workbook(excel_1)
            wb2 = load_workbook(excel_2) 
 

            # def copie_excel(excel_begin, excel_inject):
            #     for sheet in excel_inject.sheetnames:
            #         ws = excel_inject[sheet]
            #         if sheet not in excel_begin.sheetnames:
            #             excel_begin.create_sheet(title=sheet)

            #         for row in ws.iter_rows():
            #             for cell in row:
            #                 dest_cell = excel_begin[sheet][cell.coordinate]
            #                 dest_cell.font = copy(cell.font)
            #                 dest_cell.fill = copy(cell.fill)
            #                 dest_cell.border = copy(cell.border)
            #                 dest_cell.alignment = copy(cell.alignment)
            #                 dest_cell.number_format = copy(cell.number_format)
            #                 dest_cell.protection = copy(cell.protection)

            #         for row in ws.iter_rows():
            #             for cell in row:
            #                 dest_cell = excel_begin[sheet][cell.coordinate]
            #                 dest_cell.value = cell.value

            #     for sheet in excel_begin.sheetnames:
            #         ws = excel_begin[sheet]
            #         for column_cells in ws.columns:
            #             max_length = 0
            #             for cell in column_cells:
            #                 try:
            #                     if len(str(cell.value)) > max_length:
            #                         max_length = len(cell.value)
            #                 except:
            #                     pass
            #             adjusted_width = (max_length + 2) * 1.2
            #             ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width

            #     if 'Sheet' in excel_begin.sheetnames:
            #         excel_begin.remove(excel_begin['Sheet'])
            #     return excel_begin

            # if wb1:
            #    wb2 = copie_excel(wb1, wb2)
           
            
            # Ajouter la ligne 'IMPAYER'
            # sheet = wb2.active
            # last_row = sheet.max_row

            # jour_sum = sum(sheet.cell(row=row, column=2).value for row in range(2, last_row))
            # mois_sum = sum(sheet.cell(row=row, column=3).value for row in range(2, last_row))
            # annee_sum = sum(sheet.cell(row=row, column=4).value for row in range(2, last_row))

            # print(jour_sum, mois_sum, annee_sum)

          
      
           
           
            # sheet.append(['IMPAYE ou AVOIR'])

            # print(sheet)

            storage_out = f'storage/{dates.strftime("%d-%m-%Y")}/output'

            # sheet = wb1['ENCAISSEMENT']

            # Initialiser les totaux
    

         

            # # Appliquer le style de couleur verte à la ligne 'TOTAUX'
            # for cell in sheet.iter_rows(min_row=sheet.max_row, max_row=sheet.max_row):
            #     for col in cell:
            #         col.fill = PatternFill(start_color='0072BC', end_color='0072BC', fill_type='solid')
            #         col.font = Font(bold=True, color='FFFFFF')

            # for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=sheet.max_column):
            #     for cell in row:
            #         try:
            #             cell.number_format = '#,##0.00'
            #         except:
            #             pass
            # Espacer les colonnes en fonction des données
            # for col in sheet.columns:
            #     max_length = 0
            #     for cell in col:
            #         try:
            #             if len(str(cell.value)) > max_length:
            #                 max_length = len(cell.value)
            #         except:
            #             pass
            #     adjusted_width = (max_length + 2) * 1.2
            #     sheet.column_dimensions[col[0].column_letter].width = adjusted_width

            if not os.path.exists(storage_out):
                os.makedirs(storage_out)
            merged_file_path = os.path.join(storage_out,
                                            f'vente du jour{dates.strftime("%d-%m-%Y")}.xlsx')
            wb2.save(merged_file_path)

            print(f"Fichier fusionné enregistré à : {merged_file_path}")

            return merged_file_path
        except Exception as e:
            write_log(f"Erreur Exception : {str(e)}")
        finally:
            connexion.close()

    return None
