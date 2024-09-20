import os
import openpyxl
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font
from utils import get_today
import json
import pyodbc
import datetime
from utils import write_log, extract_values_in_json, get_today
from openpyxl.utils import get_column_letter



def exportDataFrameEncaissement(df, dates, header_row):
    wb = Workbook()
    ws = wb.active
    # Renommer la feuille de calcul
    ws.title = "ENCAISSEMENT"

    # Créer un dictionnaire pour stocker les données dans le format attendu
    formatted_data = {}
    for key, values in df.items():
        for value in values:
            intitule = value[0]
            amount = value[1]
            # print(intitule, value, key)

            if intitule not in formatted_data:
                formatted_data[intitule] = {'day': 0, 'month': 0, 'year': 0}
            formatted_data[intitule][key] = amount
    formatted_data['Fond de Caisse'] = {'day': 0, 'month': 200000, 'year': 200000}
    # Écrire les en-têtes des colonnes avec mise en forme
    ws.append(header_row)
    for cell in ws[1]:
        cell.font = Font(bold=True)  # Rendre l'entête en gras

    # Écrire les données dans le fichier Excel avec mise en forme
    for intitule, amounts in formatted_data.items():
        if intitule == 'Espèces':
            # print(f"Espèces : {amounts['year']} => {amounts['year'] - 200000}")
            amounts['year'] = amounts['year'] - 200000
        ws.append([intitule, amounts['day'], amounts['month'], amounts['year']])

    total_row = ['TOTAL ENCAISSEMENT', sum(amounts['day'] for amounts in formatted_data.values()),
                 sum(amounts['month'] for amounts in formatted_data.values()) - 200000,
                 sum(amounts['year'] for amounts in formatted_data.values()) - 200000]
    ws.append(total_row)

    # Formater les cellules avec des nombres au format français
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=ws.max_column):
        for cell in row:
            try:
                cell.number_format = '#,##0.00'
            except:
                pass
    # Ajouter des bordures aux cellules
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = Border(top=Side(style='thin'), bottom=Side(style='thin'),
                                 left=Side(style='thin'), right=Side(style='thin'))

    # Colorier les cellules de l'en-tête
    for cell in ws[1]:
        cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

    # Colorier la ligne TOTAUX en bleu avec texte en blanc et gras
    for row in ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.fill = PatternFill(start_color='0072BC', end_color='0072BC', fill_type='solid')
            cell.font = Font(color="FFFFFF", bold=True)

    # Espacez les colonnes selon les données
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    # Sauvegarder le fichier Excel
    storage_dir = f'storage/{dates.strftime("%d-%m-%Y")}'
    if not os.path.exists(storage_dir):
        os.makedirs(storage_dir)

    # Sauvegarder le fichier Excel dans le dossier "storage"
    file_path = os.path.join(storage_dir, 'ENCAISSEMENT.xlsx')
    wb.save(file_path)

    # Retourner le chemin absolu du fichier
    return os.path.abspath(file_path)


# def exportDataFrameSimple(df, title, cols_total):
#     if df is not None and len(df.columns) > 0:
#         # Créer un nouveau classeur Excel
#         wb = openpyxl.Workbook()
#         # Sélectionner la première feuille de calcul
#         ws = wb.active
#         ws.title = title

#         # Ajouter l'en-tête de colonne
#         for c_idx, col_name in enumerate(df.columns, start=1):
#             cell = ws.cell(row=1, column=c_idx, value=col_name)
#             cell.fill = openpyxl.styles.PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

#         # Ajouter les données du DataFrame à la feuille de calcul
#         for r_idx, row in enumerate(df.iterrows(), start=2):
#             for c_idx, value in enumerate(row[1], start=1):
#                 cell = ws.cell(row=r_idx, column=c_idx, value=value)

#         # Ajouter des bordures aux données
#         for row in ws.iter_rows(min_row=1, max_row=len(df) + 1, min_col=1, max_col=len(df.columns)):
#             for cell in row:
#                 cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),
#                                                      right=openpyxl.styles.Side(style='thin'),
#                                                      top=openpyxl.styles.Side(style='thin'),
#                                                      bottom=openpyxl.styles.Side(style='thin'))



#         # cols_total
#         col_pos = [13, 15, 16, 17, 18, 19, 20, 21]

#         # Ajouter une ligne de total à la fin pour la somme des colonnes G, H et I
#         for c_idx, _ in enumerate(df.columns, start=1):
#             if c_idx in col_pos:  # Ex: Colonnes G, H et I (indices 7, 8 et 9)
#                 total = df.iloc[:, c_idx - 1].sum()
#                 cell = ws.cell(row=len(df) + 2, column=c_idx, value=total)

#         # Ajouter le texte "TOTAUX DU JOUR" et formater la ligne
#         totaux_row = len(df) + 2  # Ligne de sous-totaux
#         ws.cell(row=totaux_row, column=12, value='TOTAUX DU JOUR').font = Font(bold=True, color='000000')
#         # for col in ws.iter_cols(min_row=totaux_row, max_row=totaux_row, min_col=1, max_col=len(df.columns)):
#         #     for cell in col:
#         #         cell.fill = openpyxl.styles.PatternFill(start_color="0072BC", end_color="0072BC", fill_type="solid")
#         #         cell.font = Font(bold=True, color='FFFFFF')


                
#         # Insert TOTAUX DU MOIS into the specified columns in the next row
#         # values_row = [13, 15, 16, 17, 18, 19, 20, 21]
#         values_row = getTotalMonth()


#         if values_row:
#             value_row_idx = len(df) + 3  # This is the row after the total row

#             for idx, c_idx in enumerate(col_pos):
#                 ws.cell(row=value_row_idx, column=c_idx, value=values_row[idx])

#             # Optionally, add a label for this row if needed
#             ws.cell(row=value_row_idx, column=12, value='TOTAUX DU MOIS').font = Font(bold=True, color='000000')



                

#         # # Ajouter une ligne de total à la fin pour la somme des colonnes D, E et F
#         # for c_idx, _ in enumerate(df.columns, start=1):
#         #     if c_idx in cols_total:  # Colonnes D, E et F (indices 3, 4 et 5)
#         #         total = df.iloc[:, c_idx - 1].sum()
#         #         cell = ws.cell(row=len(df) + 2, column=c_idx, value=total)

#         # # Ajouter le texte "TOTAUX" et formater la ligne
#         # totaux_row = len(df) + 2  # Ligne de sous-totaux
#         # ws.cell(row=totaux_row, column=1, value='TOTAUX').font = Font(bold=True, color='FFFFFF')
#         # for col in ws.iter_cols(min_row=totaux_row, max_row=totaux_row, min_col=1, max_col=len(df.columns)):
#         #     for cell in col:
#         #         cell.fill = openpyxl.styles.PatternFill(start_color="0072BC", end_color="0072BC", fill_type="solid")
#         #         cell.font = Font(bold=True, color='FFFFFF')




#         for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=ws.max_column):
#             for cell in row:
#                 try:
#                     cell.number_format = '#,##0.00'
#                 except:
#                     pass
#         # Espacer les colonnes en fonction des données
#         for col in ws.columns:
#             max_length = 0
#             for cell in col:
#                 try:
#                     if len(str(cell.value)) > max_length:
#                         max_length = len(cell.value)
#                 except:
#                     pass
#             adjusted_width = (max_length + 2) * 1.2
#             ws.column_dimensions[col[0].column_letter].width = adjusted_width

#         # Créer le dossier de stockage s'il n'existe pas
#         storage_dir = f'storage/{get_today().strftime("%d-%m-%Y")}'
#         if not os.path.exists(storage_dir):
#             os.makedirs(storage_dir)

#         # Sauvegarder le fichier Excel dans le dossier "storage"
#         file_path = os.path.join(storage_dir, f'{title}.xlsx')
#         # Enregistrer le classeur Excel
#         wb.save(file_path)
#         return os.path.abspath(file_path)

#     return None



# def exportDataFrameSimple(df, title, cols_total):
#     if df is not None and len(df.columns) > 0:
#         # Créer un nouveau classeur Excel
#         wb = openpyxl.Workbook()
#         # Sélectionner la première feuille de calcul
#         ws = wb.active
#         ws.title = title

#         # Ajouter l'en-tête de colonne pour la première feuille
#         for c_idx, col_name in enumerate(df.columns, start=1):
#             cell = ws.cell(row=1, column=c_idx, value=col_name)
#             cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
#             cell.font = Font(bold=True)

#         # Ajouter les données du DataFrame à la première feuille
#         for r_idx, row in enumerate(df.iterrows(), start=2):
#             for c_idx, value in enumerate(row[1], start=1):
#                 cell = ws.cell(row=r_idx, column=c_idx, value=value)

#         # Ajouter des bordures aux données de la première feuille
#         for row in ws.iter_rows(min_row=1, max_row=len(df) + 1, min_col=1, max_col=len(df.columns)):
#             for cell in row:
#                 cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

#         # Ajouter une ligne de total pour la première feuille
#         col_pos = [13, 15, 16, 17, 18, 19, 20, 21]
#         for c_idx, _ in enumerate(df.columns, start=1):
#             if c_idx in col_pos:
#                 total = df.iloc[:, c_idx - 1].sum()
#                 cell = ws.cell(row=len(df) + 2, column=c_idx, value=total)

#         ws.cell(row=len(df) + 2, column=12, value='TOTAUX DU JOUR').font = Font(bold=True, color='000000')


#         # Insert TOTAUX DU MOIS into the specified columns in the next row
#         values_row = getTotalMonth()

#         if values_row:
#             value_row_idx = len(df) + 3  # This is the row after the total row

#             for idx, c_idx in enumerate(col_pos):
#                 ws.cell(row=value_row_idx, column=c_idx, value=values_row[idx])

#             # Optionally, add a label for this row if needed
#             ws.cell(row=value_row_idx, column=12, value='TOTAUX DU MOIS').font = Font(bold=True, color='000000')


#         # Créer une nouvelle feuille "Detail article mois"
#         ws_month = wb.create_sheet(title="Detail article mois")
        
#         # Ajouter les en-têtes de colonne personnalisés
#         header = ["ARTICLE", "DESIGNATION", "UNITE", "QUANTITE", "QTE_TONNE"]
#         for c_idx, col_name in enumerate(header, start=1):
#             cell = ws_month.cell(row=1, column=c_idx, value=col_name)
#             cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
#             cell.font = Font(bold=True)

#         # Ajouter les lignes de détail à la nouvelle feuille
#         detail_month_rows = getDetailMonth()  # Assuming this returns the rows with 22 items each
        
#         if detail_month_rows:
#             for r_idx, row in enumerate(detail_month_rows, start=2):
#                 for c_idx, value in enumerate(row[:5], start=1):  # Write only the first 5 columns
#                     cell = ws_month.cell(row=r_idx, column=c_idx, value=value)
#                     if c_idx in [4, 5]:  # Columns D and E (index 3 and 4 in zero-based indexing)
#                         cell.number_format = '0.00'  # Format as number with two decimal places

#             # Ajouter une ligne de total pour les colonnes D et E
#             total_row_idx = len(detail_month_rows) + 2
#             total_D = sum(row[3] for row in detail_month_rows)  # Assuming QUANTITE is at index 3 (column D)
#             total_E = sum(row[4] for row in detail_month_rows)  # Assuming QTE_TONNE is at index 4 (column E)
#             ws_month.cell(row=total_row_idx, column=4, value=total_D).number_format = '0.00'  # Total for column D
#             ws_month.cell(row=total_row_idx, column=5, value=total_E).number_format = '0.00'  # Total for column E

#             # Label the total row
#             ws_month.cell(row=total_row_idx, column=3, value='TOTAUX').font = Font(bold=True)

#         # Ajouter des bordures aux données de la feuille "Detail article mois"
#         for row in ws_month.iter_rows(min_row=1, max_row=ws_month.max_row, min_col=1, max_col=5):
#             for cell in row:
#                 cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))


#         article_vente_jour = getVenteJourArticle()





#         # Créer le dossier de stockage s'il n'existe pas
#         storage_dir = f'storage/{get_today().strftime("%d-%m-%Y")}'
#         if not os.path.exists(storage_dir):
#             os.makedirs(storage_dir)

#         # Sauvegarder le fichier Excel dans le dossier "storage"
#         file_path = os.path.join(storage_dir, f'{title}.xlsx')
#         wb.save(file_path)
#         return os.path.abspath(file_path)

#     return None



def exportDataFrameSimple(df, title, turn_soc_name):
    if df is not None and len(df.columns) > 0:
        # Créer un nouveau classeur Excel
        wb = openpyxl.Workbook()
        # Sélectionner la première feuille de calcul
        ws = wb.active
        ws.title = title

        # Ajouter l'en-tête de colonne pour la première feuille
        for c_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=c_idx, value=col_name)
            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            cell.font = Font(bold=True)

        # Ajouter les données du DataFrame à la première feuille
        for r_idx, row in enumerate(df.iterrows(), start=2):
            for c_idx, value in enumerate(row[1], start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)

        # Ajouter des bordures aux données de la première feuille
        for row in ws.iter_rows(min_row=1, max_row=len(df) + 1, min_col=1, max_col=len(df.columns)):
            for cell in row:
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # # Ajouter une ligne de total pour la première feuille
        # col_pos = [13, 15, 16, 17, 18, 19, 20, 21]
        # total_row_idx = len(df) + 2

        # for c_idx in col_pos:
        #     total = df.iloc[:, c_idx - 1].sum()
        #     cell = ws.cell(row=total_row_idx, column=c_idx, value=total)
        #     cell.number_format = '#,##0.00'  # Format as number with two decimal places





        # ws.cell(row=total_row_idx, column=12, value='TOTAUX DU JOUR').font = Font(bold=True, color='000000')

        # # Ajouter des bordures aux totaux
        # for cell in ws[total_row_idx]:
        #     cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # # Insert TOTAUX DU MOIS into the specified columns in the next row
        # values_row = getTotalMonth(turn_soc_name)

        # if values_row:
        #     value_row_idx = len(df) + 3  # This is the row after the total row

        #     for idx, c_idx in enumerate(col_pos):
        #         cell = ws.cell(row=value_row_idx, column=c_idx, value=values_row[idx])
        #         cell.number_format = '#,##0.00'  # Format as number with two decimal places

        #     # Add a label for this row
        #     ws.cell(row=value_row_idx, column=12, value='TOTAUX DU MOIS').font = Font(bold=True, color='000000')

        #     # Ajouter des bordures aux totaux du mois
        #     for cell in ws[value_row_idx]:
        #         cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Espacer les colonnes en fonction des données
        for col in ws.columns:
            max_length = 0
            for cell in col:
                try:
                    if cell.value is not None and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width




        # # DETAIL ARTICLE MOIS ####################
        # # Créer une nouvelle feuille "Detail article mois"
        # ws_month = wb.create_sheet(title="Detail article mois")

        # # Ajouter les en-têtes de colonne personnalisés
        # header = ["ARTICLE", "DESIGNATION", "UNITE", "QUANTITE", "QTE_TONNE"]
        
        # for c_idx, col_name in enumerate(header, start=1):
        #     cell = ws_month.cell(row=1, column=c_idx, value=col_name)
        #     cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        #     cell.font = Font(bold=True)

        # # Ajouter les lignes de détail à la nouvelle feuille
        # detail_month_rows = getDetailMonth(turn_soc_name)  # Assuming this returns the rows with 22 items each

        # if detail_month_rows:
        #     for r_idx, row in enumerate(detail_month_rows, start=2):
        #         for c_idx, value in enumerate(row[:5], start=1):  # Write only the first 5 columns
        #             cell = ws_month.cell(row=r_idx, column=c_idx, value=value)
        #             if c_idx in [4, 5]:  # Columns D and E (index 3 and 4 in zero-based indexing)
        #                 cell.number_format = '0.00'  # Format as number with two decimal places

        #     # Ajouter une ligne de total pour les colonnes D et E
        #     total_row_idx = len(detail_month_rows) + 2
        #     total_D = sum(row[3] for row in detail_month_rows)  # Assuming QUANTITE is at index 3 (column D)
        #     total_E = sum(row[4] for row in detail_month_rows)  # Assuming QTE_TONNE is at index 4 (column E)
        #     ws_month.cell(row=total_row_idx, column=4, value=total_D).number_format = '0.00'  # Total for column D
        #     ws_month.cell(row=total_row_idx, column=5, value=total_E).number_format = '0.00'  # Total for column E

        #     # Label the total row
        #     ws_month.cell(row=total_row_idx, column=3, value='TOTAUX').font = Font(bold=True)

        # # Ajouter des bordures aux données de la feuille "Detail article mois"
        # for row in ws_month.iter_rows(min_row=1, max_row=ws_month.max_row, min_col=1, max_col=5):
        #     for cell in row:
        #         cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # # Espacer les colonnes en fonction des données pour "Detail article mois"
        # for col_idx, col in enumerate(ws_month.iter_cols(min_row=1, max_row=ws_month.max_row, min_col=1, max_col=5), start=1):
        #     max_length = 0
        #     for cell in col:
        #         try:
        #             if cell.value is not None and len(str(cell.value)) > max_length:
        #                 max_length = len(str(cell.value))
        #         except:
        #             pass
        #     adjusted_width = (max_length + 2) * 1.2
        #     ws_month.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = adjusted_width






        # # ARTICLE VENTE JOUR ####################
        # # Créer une nouvelle feuille "Article vente jour"
        # ws_vente_jour = wb.create_sheet(title="Article vente jour")

        # # Ajouter les en-têtes de colonne personnalisés
        # for c_idx, col_name in enumerate(header, start=1):
        #     cell = ws_vente_jour.cell(row=1, column=c_idx, value=col_name)
        #     cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        #     cell.font = Font(bold=True)

        # # Ajouter les lignes de détail à la nouvelle feuille "Article vente jour"
        # article_vente_jour = getVenteJourArticle(turn_soc_name)  # Assuming this returns the rows with the necessary data

        # if article_vente_jour:
        #     for r_idx, row in enumerate(article_vente_jour, start=2):
        #         for c_idx, value in enumerate(row[:5], start=1):  # Write only the first 5 columns
        #             cell = ws_vente_jour.cell(row=r_idx, column=c_idx, value=value)
        #             if c_idx in [4, 5]:  # Columns D and E (index 3 and 4 in zero-based indexing)
        #                 cell.number_format = '0.00'  # Format as number with two decimal places

        #     # Ajouter une ligne de total pour les colonnes D et E
        #     total_row_idx = len(article_vente_jour) + 2
        #     total_D = sum(row[3] for row in article_vente_jour)  # Assuming QUANTITE is at index 3 (column D)
        #     total_E = sum(row[4] for row in article_vente_jour)  # Assuming QTE_TONNE is at index 4 (column E)
        #     ws_vente_jour.cell(row=total_row_idx, column=4, value=total_D).number_format = '0.00'  # Total for column D
        #     ws_vente_jour.cell(row=total_row_idx, column=5, value=total_E).number_format = '0.00'  # Total for column E

        #     # Label the total row
        #     ws_vente_jour.cell(row=total_row_idx, column=3, value='TOTAUX').font = Font(bold=True)

        # # Ajouter des bordures aux données de la feuille "Article vente jour"
        # for row in ws_vente_jour.iter_rows(min_row=1, max_row=ws_vente_jour.max_row, min_col=1, max_col=5):
        #     for cell in row:
        #         cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # # Espacer les colonnes en fonction des données
        # for col in ws_vente_jour.columns:
        #     max_length = 0
        #     for cell in col:
        #         try:
        #             if len(str(cell.value)) > max_length:
        #                 max_length = len(cell.value)
        #         except:
        #             pass
        #     adjusted_width = (max_length + 2) * 1.2
        #     ws_vente_jour.column_dimensions[col[0].column_letter].width = adjusted_width








        # # CATEGORIE PRODUIT ####################
        # # Créer une nouvelle feuille "Catégorie produit"
        # categorisation_produit = wb.create_sheet(title="Vente Jour Catégorie")

        # # Ajouter les en-têtes de colonne personnalisés
        # header = ["FAMILLE", "INTITULE", "UNITE", "QUANTITE", "QTE_TONNE"]
        
        # for c_idx, col_name in enumerate(header, start=1):
        #     cell = categorisation_produit.cell(row=1, column=c_idx, value=col_name)
        #     cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        #     cell.font = Font(bold=True)

        # # Ajouter les lignes de détail à la nouvelle feuille "Catégorie produit"
        # article_vente_categorie = getCategorieArticle(turn_soc_name)  # Assuming this returns the rows with the necessary data

        # if article_vente_categorie:
        #     for r_idx, row in enumerate(article_vente_categorie, start=2):
        #         for c_idx, value in enumerate(row[:5], start=1):  # Write only the first 5 columns
        #             cell = categorisation_produit.cell(row=r_idx, column=c_idx, value=value)
        #             if c_idx in [4, 5]:  # Columns D and E (index 3 and 4 in zero-based indexing)
        #                 cell.number_format = '0.00'  # Format as number with two decimal places

        #     # Ajouter une ligne de total pour les colonnes D et E
        #     total_row_idx = len(article_vente_categorie) + 2
        #     total_D = sum(row[3] for row in article_vente_categorie)  # Assuming QUANTITE is at index 3 (column D)
        #     total_E = sum(row[4] for row in article_vente_categorie)  # Assuming QTE_TONNE is at index 4 (column E)
        #     categorisation_produit.cell(row=total_row_idx, column=4, value=total_D).number_format = '0.00'  # Total for column D
        #     categorisation_produit.cell(row=total_row_idx, column=5, value=total_E).number_format = '0.00'  # Total for column E

        #     # Label the total row
        #     categorisation_produit.cell(row=total_row_idx, column=3, value='TOTAUX').font = Font(bold=True)

        # # Ajouter des bordures aux données de la feuille "Catégorie produit"
        # for row in categorisation_produit.iter_rows(min_row=1, max_row=categorisation_produit.max_row, min_col=1, max_col=5):
        #     for cell in row:
        #         cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # # Espacer les colonnes en fonction des données
        # for col in categorisation_produit.columns:
        #     max_length = 0
        #     for cell in col:
        #         try:
        #             if len(str(cell.value)) > max_length:
        #                 max_length = len(cell.value)
        #         except:
        #             pass
        #     adjusted_width = (max_length + 2) * 1.2
        #     categorisation_produit.column_dimensions[col[0].column_letter].width = adjusted_width






        #################### FINAL BUILD ####################

        # Créer le dossier de stockage s'il n'existe pas
        storage_dir = f'storage/{get_today().strftime("%d-%m-%Y")}'
        if not os.path.exists(storage_dir):
            os.makedirs(storage_dir)

        # Sauvegarder le fichier Excel dans le dossier "storage"
        file_path = os.path.join(storage_dir, f'{title}.xlsx')
        wb.save(file_path)
        return os.path.abspath(file_path)

    return None



  
def getCategorieArticle(turn_soc_name):
    with open('societe.json', 'r', encoding='utf-8') as file:
        societe_json = json.load(file)

    for soc in societe_json:
        if soc['nom'] == turn_soc_name:

            server = soc['server']
            base = soc['base']

            cnxn = connexionSQlServer(server, base, soc_name=None)

            if cnxn:
                cursor = cnxn.cursor()

                # Get today's date formatted as YYYY-MM-DD
                today = datetime.datetime.today()
                date_end = today.strftime('%Y-%d-%m')

                # Get the first day of the current month formatted as YYYY-MM-DD
                date_start = today.replace(day=1).strftime('%Y-%d-%m')


                # getting all the requests
                requests = extract_values_in_json('SQL')

                # choosing the query:
                for key, value in requests.items():
                    # if key == "totalMoisX3":

                    querryType = getCategorieArticleType(soc['type'])

                    if key == querryType:

                        # cursor.execute(str(value).replace("{ag_vets}", "AGVAL"), (date_start, date_end))
                        # cursor.execute(str(value).replace("{ag_vets}", "AGVAL"))
                        query = str(value).replace("{ag_vets}", soc['valeur']).replace("{today}", today.strftime('%Y-%d-%m'))
                        cursor.execute(query)


                        # Fetch the row containing the 8 sums
                        values_row = cursor.fetchall()

                        # print(values_row)

                        # Return the row as an array of 8 values
                        if values_row:
                            return values_row
                        
                        break
                
    return None



def getDetailMonth(turn_soc_name):
    with open('societe.json', 'r', encoding='utf-8') as file:
        societe_json = json.load(file)

    for soc in societe_json:
        if soc['nom'] == turn_soc_name:
            server = soc['server']
            base = soc['base']

            cnxn = connexionSQlServer(server, base, soc_name=None)

            if cnxn:
                cursor = cnxn.cursor()

                # Get today's date formatted as YYYY-MM-DD
                today = datetime.datetime.today()
                date_end = today.strftime('%Y-%d-%m')

                # Get the first day of the current month formatted as YYYY-MM-DD
                date_start = today.replace(day=1).strftime('%Y-%d-%m')


                # getting all the requests
                requests = extract_values_in_json('SQL')

                # choosing the query:
                for key, value in requests.items():
                    # if key == "totalMoisX3":

                    querryType = getDetailArtMoisType(soc['type'])

                    if key == querryType:

                        # cursor.execute(str(value).replace("{ag_vets}", "AGVAL"), (date_start, date_end))
                        # cursor.execute(str(value).replace("{ag_vets}", "AGVAL"))
                        query = str(value).replace("{ag_vets}", soc['valeur'])
                        cursor.execute(query)


                        # Fetch the row containing the 8 sums
                        values_row = cursor.fetchall()

                        # print(values_row)

                        # Return the row as an array of 8 values
                        if values_row:
                            return values_row
                        
                        break
                    
    return None

 

def getVenteJourArticle(turn_soc_name):
    with open('societe.json', 'r', encoding='utf-8') as file:
        societe_json = json.load(file)

    for soc in societe_json:
        if soc['nom'] == turn_soc_name:
            server = soc['server']
            base = soc['base']

            cnxn = connexionSQlServer(server, base, soc_name=None)

            if cnxn:
                cursor = cnxn.cursor()

                # Get today's date formatted as YYYY-MM-DD
                today = datetime.datetime.today()
                date_end = today.strftime('%Y-%d-%m')

                # Get the first day of the current month formatted as YYYY-MM-DD
                date_start = today.replace(day=1).strftime('%Y-%d-%m')


                # getting all the requests
                requests = extract_values_in_json('SQL')

                # choosing the query:
                for key, value in requests.items():
                    # if key == "totalMoisX3":

                    querryType = getVentJourArticleType(soc['type'])

                    if key == querryType:

                        # cursor.execute(str(value).replace("{ag_vets}", "AGVAL"), (date_start, date_end))
                        # cursor.execute(str(value).replace("{ag_vets}", "AGVAL"))
                        query = str(value).replace("{ag_vets}", soc['valeur']).replace("{today}", today.strftime('%Y-%d-%m'))
                        cursor.execute(query)


                        # Fetch the row containing the 8 sums
                        values_row = cursor.fetchall()

                        # print(values_row)

                        # Return the row as an array of 8 values
                        if values_row:
                            return values_row
                        
                        break
                    
    return None

 


def getTotalMonth(turn_soc_name):

    with open('societe.json', 'r', encoding='utf-8') as file:
        societe_json = json.load(file)

    for soc in societe_json:
        if soc['nom'] == turn_soc_name:
            server = soc['server']
            base = soc['base']

            cnxn = connexionSQlServer(server, base, soc_name=None)

            if cnxn:
                cursor = cnxn.cursor()

                # # Get today's date formatted as YYYY-MM-DD
                # today = datetime.datetime.today()
                # date_end = today.strftime('%Y-%d-%m')

                # # Get the first day of the current month formatted as YYYY-MM-DD
                # date_start = today.replace(day=1).strftime('%Y-%d-%m')


                # getting all the requests
                requests = extract_values_in_json('SQL')

                # choosing the query:
                for key, value in requests.items():

                    querryType = getTotalSocietyType(soc['type'])

                    if key == querryType:
                    
                    # if key == "x3":

                        # cursor.execute(str(value).replace("{ag_vets}", "AGVAL"), (date_start, date_end))
                        cursor.execute(str(value).replace("{ag_vets}", soc['valeur']))
                        # query = str(value).replace("{today}", today.strftime('%Y-%d-%m')).replace("{ag_vets}", "AGVAL")
                        # cursor.execute(query)


                        # Fetch the row containing the 8 sums
                        values_row = cursor.fetchone()

                        # print(values_row)

                        # Return the row as an array of 8 values
                        if values_row:
                            return list(values_row)
                        
                        break
                    
    return None

 

# class Total_month:
#     def __init__(self):
#         self.Quantite = None
#         self.unit = None
#         self.Qte_tonne = None
#         self.Prix_unitaire = None
#         self.Prix_net = None
#         self.Prix_TTC = None
#         self.Remise = None
#         self.Montant_HT = None
#         self.Montant_TT = None



def connexionSQlServer(server, base, soc_name=None):
    conn = None
    value = f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={server};DATABASE={base};UID=reader;PWD=m1234"
    # print(value)
    try:
        conn = pyodbc.connect(value)
    except pyodbc.Error as e:
        # Si une erreur se produit lors de la connexion, ajoutez le serveur à la liste des bases sans accès
        print("===============================")
        print(f"Erreur de Connexion pour {server} à la base {soc_name} ")
        print("===============================")
        write_log(f"Erreur de connexion : {str(e)}")
    except Exception as e:
        write_log(f"Erreur Exception : {str(e)}")
    return conn
    



def merge_excel_files(excel_1_path, excel_2_path, merged_file_path):
    # Charger les fichiers Excel existants
    wb1 = load_workbook(filename=excel_1_path)
    wb2 = load_workbook(filename=excel_2_path)

    # Créer un nouveau classeur Excel pour la fusion
    merged_wb = load_workbook()

    # Copier les feuilles de excel_1 dans le classeur fusionné
    for sheet_name in wb1.sheetnames:
        source = wb1[sheet_name]
        target = merged_wb.create_sheet(sheet_name)
        for row in source.iter_rows(min_row=1, max_row=source.max_row, min_col=1, max_col=source.max_column):
            for cell in row:
                target[cell.coordinate].value = cell.value

    # Copier les feuilles de excel_2 dans le classeur fusionné
    for sheet_name in wb2.sheetnames:
        source = wb2[sheet_name]
        target = merged_wb.create_sheet(sheet_name)
        for row in source.iter_rows(min_row=1, max_row=source.max_row, min_col=1, max_col=source.max_column):
            for cell in row:
                target[cell.coordinate].value = cell.value

    # Supprimer la feuille par défaut créée par load_workbook()
    merged_wb.remove(merged_wb['Sheet'])

    # Enregistrer le classeur fusionné
    merged_wb.save(filename=merged_file_path)

    return merged_file_path



def get_vet_name(soc_name):
    if "AGRIVET" in soc_name:
        return "AGVET"
    elif "AGRIVAL" in soc_name:
        return "AGVAL"
    elif "AGRIKOBA" in soc_name:
        return "AGKOB"
    elif "AGRIFARM" in soc_name:
        return "AGFRM"
    else:
        return "SMTP"


def getTotalSocietyType(type):
    if type == "sage100":
        return "totalMoisSage100"
    elif type == "x3":
        return "totalMoisX3"
    elif type == "smtpx3":
        return "totalMoisSmtpX3"
    return None
    

def getDetailJourSocietyType(type):
    if type == "sage100":
        return "detailJourSage100"
    elif type == "x3":
        return "detailJourX3"
    elif type == "smtpx3":
        return "detailJourSmtpX3"
    return None
    
    

def getDetailArtMoisType(type):
    if type == "sage100":
        return "detailArtMoisSage100"
    elif type == "x3":
        return "detailArtMoisX3"
    elif type == "smtpx3":
        return "detailArtMoisSmtpX3"
    return None
    
def getVentJourArticleType(type):
    if type == "sage100":
        return "venteJourArticleSage100"
    elif type == "x3":
        return "venteJourArticleX3"
    elif type == "smtpx3":
        return "venteJourArticleSmtpX3"
    return None
        
def getCategorieArticleType(type):
    if type == "sage100":
        return "categorieArticleSage100"
    elif type == "x3":
        return "categorieArticleX3"
    elif type == "smtpx3":
        return "categorieArticleSmtpX3"
    return None
    