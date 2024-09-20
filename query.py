import os
from copy import copy

import pyodbc
import pandas as pd
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill, Font

from export import exportDataFrameEncaissement, exportDataFrameSimple, merge_excel_files
from utils import write_log, extract_values_in_json, get_today
import decimal
from datetime import datetime


def connexionSQlServer(server, base, soc_name=None):
    conn = None
    value = f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={server};DATABASE={base};UID=reader;PWD=m1234"
    # print(value)
    try:
        conn = pyodbc.connect(value)
    except pyodbc.Error as e:
        # Si une erreur se produit lors de la connexion, ajoutez le serveur à la liste des bases sans accès
        print("===============================")
        print(f"Erreur de Connexion pour {server} à la base {soc_name} ")
        print("===============================")
        write_log(f"Erreur de connexion : {str(e)}")
    except Exception as e:
        write_log(f"Erreur Exception : {str(e)}")
    return conn


def get_column_index_by_title(sheet, title):
    for col in sheet.iter_cols(1, sheet.max_column):
        if col[0].value == title:
            return col[0].col_idx
    return 3


    

    
def getDataLink(connexion, dates=None, soc_name=None, base_type=None):
    if connexion is not None:
        try:
            # getting all the requests
            requests = extract_values_in_json('SQL')
            if dates is None:
                dates = get_today()
            details = None

          
            # key is "details" and values is "request" in json file section SQL
            # choosing the query:
            for key, value in requests.items():



                if base_type == key and key == "base":
                    with connexion.cursor() as cursor:
                        query = str(value).replace("{today}", dates.strftime('%Y-%m-%d'))
                        cursor.execute(query)
                        rows = cursor.fetchall()
                        rows = [tuple(row) for row in rows]
                        if all(isinstance(row, tuple) for row in rows):
                            details = pd.DataFrame(rows,
                                                   columns=["DATE", "Action", "SOURCE", "DESTINATION", "DB", "HIER", "STATUS", "PATH"])
                            




                        else:
                            print("Pas un tuple")






            excel_2 = exportDataFrameSimple(details, f"VENTE DU JOUR DETAIL {dates.strftime('%d-%m-%Y')}",
                                            soc_name)
     

            storage_dir = f'storage/{dates.strftime("%d-%m-%Y")}'
            if not os.path.exists(storage_dir):
                os.makedirs(storage_dir)

            wb2 = load_workbook(excel_2) 
 
            storage_out = f'storage/{dates.strftime("%d-%m-%Y")}/output'


            if not os.path.exists(storage_out):
                os.makedirs(storage_out)
            merged_file_path = os.path.join(storage_out,
                                            f'BASE_{soc_name.replace(" ", "_")}_{dates.strftime("%d-%m-%Y")}.xlsx')
            wb2.save(merged_file_path)

            print(f"Fichier fusionné enregistré à : {merged_file_path}")

            return merged_file_path
        except Exception as e:
            write_log(f"Erreur Exception : {str(e)}")
        finally:
            connexion.close()

    return None

    
def get_vet_name(soc_name):
    if "AGRIVET" in soc_name:
        return "AGVET"
    elif "AGRIVAL" in soc_name:
        return "AGVAL"
    elif "AGRIKOBA" in soc_name:
        return "AGKOB"
    elif "AGRIFARM" in soc_name:
        return "AGFRM"
    else:
        return "SMTP"
